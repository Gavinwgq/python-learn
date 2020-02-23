from bs4 import BeautifulSoup
import requests
import openpyxl

# 从网页爬取数据，并保存到excel中
def getCityLis(url):
    html = requests.get(url).text
    soup = BeautifulSoup(html, features='html.parser')
    return soup.find_all(class_='dropdown-menu')[0].children


def getDataList(url):
    html = requests.get(url).text
    soup = BeautifulSoup(html, features='html.parser')
    return soup.find_all(class_='clearfix brt_li')


cityList = getCityLis("http://www.itdp-china.org/brt/city/?city_id=43&city_name=%E5%AE%9C%E6%98%8C&lang=0")
urlP = 'http://www.itdp-china.org'
f = openpyxl.Workbook()
for city in cityList:
    if city == '\n':
        continue
    name = city.a.get_text().strip().replace("\n", "").replace('\r', '')
    href = city.a['href']
    print("------------------------------------")
    print(name, ':', href)
    dataList = getDataList(urlP + href)
    sheet1 = f.create_sheet(title=name)
    jk = 0
    for item in dataList:
        jk = jk + 1
        key = item.a.get_text().strip().replace("\n", "").replace('\r', '')
        if item.img is not None:
            src = item.img['src']
            if "/media/bike/img/tick.gif" == src:
                val = '是'
            if "/media/bike/img/cross.gif" == src:
                val = "否"
            if "/media/bike/img/partial.gif" == src:
                val = "部分"
        if item.span is not None:
            val = item.span.get_text().strip().replace("\n", "").replace('\r', '')
        print(key, val)
        sheet1.cell(row=jk, column=1).value = key
        sheet1.cell(row=jk, column=2).value = val
f.save("chatPy.xlsx")
