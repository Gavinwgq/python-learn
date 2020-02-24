import mysql.connector
import openpyxl
from openpyxl.styles import Font

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  passwd="root",
  database="test"
)

mycursor = mydb.cursor()

mycursor.execute("SELECT DISTINCT itemKey,title FROM brtData ORDER BY title")

myresult = mycursor.fetchall()

mycursor.execute("SELECT DISTINCT city FROM brtData")

cityList = mycursor.fetchall()

f = openpyxl.Workbook()
sheet1 = f.create_sheet(title='data')
jk = 1
sheet1.cell(row=jk, column=1).value = 'title'
sheet1.cell(row=jk, column=2).value = 'itemKey'
cn = 3
for city in cityList:
  sheet1.cell(row=jk, column=cn).value = city[0]
  sheet1.cell(row=jk, column=cn).font = Font(size=14, color='3CB371')
  cn = cn+1
for x in myresult:
  jk = jk+1
  sheet1.cell(row=jk, column=1).value = x[1]
  sheet1.cell(row=jk, column=2).value = x[0]

jk = 1
for x in myresult:
  jk = jk+1
  cn = 3
  for city in cityList:
    sql = 'SELECT itemVal from brtData where city = %s and title = %s and itemKey = %s'
    param = (city[0], x[1], x[0])
    mycursor.execute(sql,param)
    result = mycursor.fetchall()
    if len(result)>0:
      sheet1.cell(row=jk, column=cn).value = result[0][0]
    else:
      sheet1.cell(row=jk, column=cn).value = "-----"
    cn = cn+1
f.save("brtData0218.xlsx")